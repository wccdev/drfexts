import codecs
from typing import Any, Optional

import orjson
import unicodecsv as csv
from django.conf import settings
from openpyxl import load_workbook
from rest_framework.parsers import BaseParser, ParseError

__all__ = ["CustomJSONParser", "CustomXLSXParser", "CustomCSVParser"]


class CustomJSONParser(BaseParser):
    """
    Parses JSON-serialized data by orjson parser.
    """

    media_type: str = "application/json"

    def parse(
        self,
        stream,
        media_type: Optional[str] = None,
        parser_context: Any = None,
    ) -> Any:
        """
        De-serializes JSON strings to Python objects.
        :param stream: A stream-like object representing the body of the request.
        :param media_type: If provided, this is the media type of the incoming
                request content specified in the `Content-Type` HTTP header.
        :param parser_context: If supplied, this argument will be a dictionary
                containing any additional context that may be required to parse
                the request content.
                By default this will include the following
                keys: view, request, args, kwargs.
        :return: Python native instance of the JSON string.
        """
        parser_context = parser_context or {}
        encoding: str = parser_context.get("encoding", settings.DEFAULT_CHARSET)

        try:
            decoded_stream = codecs.getreader(encoding)(stream)
            return orjson.loads(decoded_stream.read())
        except ValueError as exc:
            raise ParseError("JSON parse error - %s" % str(exc))


class CustomXLSXParser(BaseParser):
    """
    Parses data frame from Excel (.xlsx)
    """

    media_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def parse(self, stream, media_type=None, parser_context=None):
        """
        Simply return a string representing the body of the request.
        """
        try:
            workbook = load_workbook(stream._stream, read_only=True)
            sheet = workbook.active
            data = []
            headers = None
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() for h in row]
                    continue
                # 构造每一行的字典
                row_dict = dict(zip(headers, row))
                data.append(row_dict)

            return data
        except Exception as exc:
            raise ParseError("Excel parse error - %s" % str(exc))
        finally:
            workbook.close()


class CustomCSVParser(BaseParser):
    """
    Parses data frame from CSV (.csv)
    """

    media_type: str = "text/csv"

    def parse(self, stream, media_type=None, parser_context=None):
        parser_context = parser_context or {}
        delimiter = parser_context.get("delimiter", ",")
        encoding = parser_context.get("encoding", settings.DEFAULT_CHARSET)

        try:
            strdata = stream.read()
            binary = self.universal_newlines(strdata)
            rows = self.unicode_csv_reader(binary, delimiter=delimiter, charset=encoding)
            data = []
            headers = next(rows)
            for row in rows:
                row_data = dict(zip(headers, row))
                data.append(row_data)
            return data
        except Exception as exc:
            raise ParseError("CSV parse error - %s" % str(exc))

    def unicode_csv_reader(self, csv_data, dialect=csv.excel, charset="utf-8", **kwargs):
        csv_reader = csv.reader(csv_data, dialect=dialect, encoding=charset, **kwargs)
        for row in csv_reader:
            yield row

    def universal_newlines(self, stream):
        # It's possible that the stream was not opened in universal
        # newline mode. If not, we may have a single "row" that has a
        # bunch of carriage return (\r) characters that should act as
        # newlines. For that case, lets call splitlines on the row. If
        # it doesn't have any newlines, it will return a list of just
        # the row itself.
        for line in stream.splitlines():
            yield line
